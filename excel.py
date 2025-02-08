import openpyxl

owb=openpyxl.Workbook()
filename='myfile/myels.xlsx'
sheet=owb.active
sheet.title='mytile'

plama=[
    ["類別","編號","品名","單價","單位"],
    ["蔬菜","01","高麗菜","40","粒"],
    ["水果","02","草莓","500","箱"],
    ["堅果","03","杏仁","150","罐"]

]

for i in range(0,4):
    for j in range(0,len(plama[i])):
        sheet.cell(row=i+1,column=j+1,value=plama[i][j])
owb.save(filename)
print("done")

owb=openpyxl.load_workbook(filename)
sheet=owb.get_sheet_by_name('mytile')

for r in sheet.rows:
    for c in r:
        print(c.value,"\t" ,end="")
        print()